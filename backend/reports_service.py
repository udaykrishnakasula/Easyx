"""Admin reporting / export service.

Builds tabular datasets (headers + rows) for the 8 exportable entity types and
renders them to CSV or XLSX. All access is admin-only (enforced at the router);
this module contains no auth logic itself.
"""
import csv
import io

from bson.decimal128 import Decimal128

from db import db
from money import fmt

# Datasets the admin can export.
DATASETS = [
    "users",
    "deposits",
    "investments",
    "matured_investments",
    "withdrawals",
    "referral_commissions",
    "wallet_transactions",
    "kyc",
]


def _num(v):
    """Money/decimal -> plain 2dp string; blank for None."""
    if v is None:
        return ""
    try:
        return fmt(v)
    except Exception:  # noqa: BLE001
        return str(v)


def _s(v):
    """Safe string; Decimal128 -> plain number; None -> blank."""
    if v is None:
        return ""
    if isinstance(v, Decimal128):
        return fmt(v)
    return str(v)


async def _users_map() -> dict:
    """user_id -> {name,email} for enriching money datasets."""
    m = {}
    async for u in db.users.find({}, {"id": 1, "name": 1, "email": 1}):
        m[u["id"]] = {"name": u.get("name", ""), "email": u.get("email", "")}
    return m


async def build_dataset(dataset: str) -> tuple[list[str], list[list]]:
    """Return (headers, rows) for the requested dataset."""
    if dataset == "users":
        headers = ["id", "name", "email", "phone", "role", "status",
                   "email_verified", "referral_code", "referred_by", "created_at"]
        rows = []
        async for u in db.users.find({}).sort("created_at", -1):
            rows.append([
                _s(u.get("id")), _s(u.get("name")), _s(u.get("email")), _s(u.get("phone")),
                _s(u.get("role")), _s(u.get("status")), _s(u.get("email_verified")),
                _s(u.get("referral_code")), _s(u.get("referred_by")), _s(u.get("created_at")),
            ])
        return headers, rows

    if dataset in ("deposits",):
        umap = await _users_map()
        headers = ["id", "user_id", "user_email", "network", "amount", "approved_amount",
                   "status", "tx_hash", "created_at", "reviewed_at"]
        rows = []
        async for d in db.deposits.find({}).sort("created_at", -1):
            u = umap.get(d.get("user_id"), {})
            rows.append([
                _s(d.get("id")), _s(d.get("user_id")), _s(u.get("email")), _s(d.get("network")),
                _num(d.get("amount")), _num(d.get("approved_amount")), _s(d.get("status")),
                _s(d.get("tx_hash")), _s(d.get("created_at")), _s(d.get("reviewed_at") or d.get("updated_at")),
            ])
        return headers, rows

    if dataset in ("investments", "matured_investments"):
        umap = await _users_map()
        query = {"status": "matured"} if dataset == "matured_investments" else {"status": {"$ne": "pending"}}
        headers = ["id", "user_id", "user_email", "plan_key", "principal", "profit_amount",
                   "maturity_amount", "profit_percentage", "status", "start_at", "maturity_at",
                   "matured_at", "created_at"]
        rows = []
        async for i in db.investments.find(query).sort("created_at", -1):
            u = umap.get(i.get("user_id"), {})
            rows.append([
                _s(i.get("id")), _s(i.get("user_id")), _s(u.get("email")), _s(i.get("plan_key")),
                _num(i.get("principal")), _num(i.get("profit_amount")), _num(i.get("maturity_amount")),
                _num(i.get("profit_percentage_snapshot")), _s(i.get("status")),
                _s(i.get("start_at")), _s(i.get("maturity_at")), _s(i.get("matured_at")), _s(i.get("created_at")),
            ])
        return headers, rows

    if dataset == "withdrawals":
        umap = await _users_map()
        headers = ["id", "user_id", "user_email", "network", "amount", "to_address",
                   "status", "tx_hash", "created_at", "processed_at"]
        rows = []
        async for w in db.withdrawals.find({}).sort("created_at", -1):
            u = umap.get(w.get("user_id"), {})
            rows.append([
                _s(w.get("id")), _s(w.get("user_id")), _s(u.get("email")), _s(w.get("network")),
                _num(w.get("amount")), _s(w.get("to_address")), _s(w.get("status")),
                _s(w.get("tx_hash")), _s(w.get("created_at")), _s(w.get("processed_at") or w.get("updated_at")),
            ])
        return headers, rows

    if dataset == "referral_commissions":
        umap = await _users_map()
        headers = ["id", "referrer_id", "referrer_email", "referee_id", "referee_email",
                   "investment_id", "plan_key", "amount", "percentage", "status", "created_at"]
        rows = []
        async for c in db.referral_commissions.find({}).sort("created_at", -1):
            rr = umap.get(c.get("referrer_id"), {})
            re = umap.get(c.get("referee_id"), {})
            rows.append([
                _s(c.get("id")), _s(c.get("referrer_id")), _s(rr.get("email")),
                _s(c.get("referee_id")), _s(re.get("email")), _s(c.get("investment_id")),
                _s(c.get("plan_key")), _num(c.get("amount")), _num(c.get("percentage")),
                _s(c.get("status")), _s(c.get("created_at")),
            ])
        return headers, rows

    if dataset == "wallet_transactions":
        umap = await _users_map()
        headers = ["id", "user_id", "user_email", "type", "direction", "amount",
                   "balance_after", "ref_type", "ref_id", "note", "created_at"]
        rows = []
        async for t in db.wallet_transactions.find({}).sort("created_at", -1):
            u = umap.get(t.get("user_id"), {})
            rows.append([
                _s(t.get("id")), _s(t.get("user_id")), _s(u.get("email")), _s(t.get("type")),
                _s(t.get("direction")), _num(t.get("amount")), _num(t.get("balance_after")),
                _s(t.get("ref_type")), _s(t.get("ref_id")), _s(t.get("note")), _s(t.get("created_at")),
            ])
        return headers, rows

    if dataset == "kyc":
        umap = await _users_map()
        headers = ["id", "user_id", "user_email", "status", "id_type",
                   "submitted_at", "reviewed_at", "reason"]
        rows = []
        async for k in db.kyc_records.find({}).sort("submitted_at", -1):
            u = umap.get(k.get("user_id"), {})
            rows.append([
                _s(k.get("id")), _s(k.get("user_id")), _s(u.get("email")), _s(k.get("status")),
                _s(k.get("id_type")), _s(k.get("submitted_at")),
                _s(k.get("reviewed_at")), _s(k.get("rejection_reason") or k.get("reason")),
            ])
        return headers, rows

    raise ValueError(f"Unknown dataset: {dataset}")


def to_csv_bytes(headers: list[str], rows: list[list]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")  # BOM => Excel opens UTF-8 correctly


def to_xlsx_bytes(headers: list[str], rows: list[list], sheet_name: str = "Report") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Report"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append(r)
    ws.freeze_panes = "A2"
    # Reasonable column widths.
    for idx, h in enumerate(headers, start=1):
        col = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[col].width = min(max(len(str(h)) + 2, 12), 40)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
